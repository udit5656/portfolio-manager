# -*- coding: utf-8 -*-
"""
Created on Tue Apr 30 22:17:31 2024

@author: Udit
"""

from openpyxl import load_workbook
import os
import json
from json import JSONEncoder
from listedCompanies import Exchanges
from Constants import constant
from dataClass import TransactionType


def get_working_dir():
    return os.getcwd()


class GrowwReport:
    exchanges = Exchanges()

    def __init__(self):
        self.transactions = []
        self.sheet = load_workbook(filename=constant.GROWW_REPORT).active
        self.read_report()

    def add(self, transaction):
        self.transactions.append(transaction)

    def read_report(self):
        self.read_trades(constant.REALISED_TRADES)
        self.read_trades(constant.UNREALISED_TRADES)
        print(len(self.transactions))

    def read_trades(self, trade_type):
        first_cell = self.get_first_trans(trade_type) + 3
        for transaction in self.sheet.iter_rows(
            min_row=first_cell,
            max_row=first_cell + constant.MAX_TRANSACTIONS,
            min_col=1,
            max_col=10,
            values_only=True,
        ):
            if transaction[0] is None:
                break

            self.add_buy_transaction(
                transaction[0],
                transaction[1],
                transaction[2],
                transaction[3],
                transaction[4],
            )

            if trade_type == constant.REALISED_TRADES:
                self.add_sell_transaction(
                    transaction[0],
                    transaction[1],
                    transaction[2],
                    transaction[6],
                    transaction[7],
                )

    def add_buy_transaction(self, stock_name, ISIN, quantity, date, price):
        self.transactions.append(
            GrowwTransaction(
                TransactionType.BUY, stock_name, ISIN, quantity, date, price
            )
        )

    def add_sell_transaction(self, stock_name, ISIN, quantity, date, price):
        self.transactions.append(
            GrowwTransaction(
                TransactionType.SELL, stock_name, ISIN, quantity, date, price
            )
        )

    def save_transactions(self):
        json_transactions = []
        for transaction in self.transactions:
            json_transactions.append(GrowwTransactionEncoder().encode(transaction))
        json_file = open(constant.JSON_FILE, "w")
        json.dump(json_transactions, json_file, indent=4, ensure_ascii=False)
        json_file.close()

    def get_first_trans(self, title):
        for idx in range(1, 1000):
            cell_value = self.sheet[f"A{idx}"].value
            if cell_value == title:
                return idx
        raise Exception(f"Cell with {title}")


class GrowwTransaction:

    def __init__(
        self,
        transaction_type,
        stock_name,
        ISIN,
        quantity,
        date,
        price,
    ):
        self.type = transaction_type
        self.stockName = stock_name
        self.ISIN = ISIN
        self.quantity = quantity
        self.date = date
        self.price = price
        self.symbol = GrowwReport.exchanges.find_symbol(self.ISIN)
        print(f"{stock_name} : {self.symbol}")


class GrowwTransactionEncoder(JSONEncoder):
    def default(self, o):
        return o.__dict__
